"""
excel_to_oscal.py — Workshop Edition (Discovery-Driven)

Converts a FedRAMP Moderate template SSP (Excel) into OSCAL JSON.
Reads what the SSP says — extracts components mentioned in narratives
via keyword matching. Does NOT assume what tools should exist or
what controls they map to. Discovery fills that in.

USAGE:
  python excel_to_oscal.py --input templates/fedramp-moderate-template-ssp.xlsx --output oscal/
  python excel_to_oscal.py --help

MODELS PRODUCED:
  Model 4 (SSP):          oscal/ssp.json — the claim
  Model 6 (AR):           oscal/assessment-results.json — skeleton
  Model 7 (POA&M):        oscal/poam.json — skeleton
"""

import json
import os
import sys
import re
import uuid
import argparse
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Run: pip install openpyxl")
    sys.exit(1)

# Import shared utilities
sys.path.insert(0, os.path.dirname(__file__))
from pipeline_utils import (
    stable_uuid, OSCAL_NAMESPACE, KNOWN_COMPONENTS, extract_components_from_text,
    oscal_roles_and_parties,
)


# ── Excel configuration ───────────────────────────────────────────────────────

SHEET_NAME = "FedRAMP Moderate Baseline"
HEADER_ROW = 1
DATA_START_ROW = 2

COLUMNS = {
    "number":          1,
    "family":          2,
    "control_id":      3,
    "control_name":    4,
    "control_text":    5,
    "related":         6,
    "status":          7,
    "evidence_method": 8,
    "implementation":  9,
}

VALID_STATUSES = {
    "implemented":    "implemented",
    "inherited":      "inherited",
    "planned":        "planned",
    "not applicable": "not-applicable",
    "not-applicable": "not-applicable",
    "n/a":            "not-applicable",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_control_id(raw_id: str) -> str:
    if not raw_id:
        return None
    s = str(raw_id).strip().lower()
    s = re.sub(r'-0*(\d)', r'-\1', s)
    s = re.sub(r'\(0*(\d+)\)', r'(\1)', s)
    return s


def normalize_status(raw_status: str) -> str:
    if not raw_status:
        return "not-implemented"
    key = str(raw_status).strip().lower()
    return VALID_STATUSES.get(key, "not-implemented")


def get_cell(sheet, row: int, col: int):
    val = sheet.cell(row=row, column=col).value
    if val is None:
        return None
    cleaned = str(val).strip()
    return cleaned if cleaned else None


def is_missing_or_stale(text, status):
    if not text:
        return True, "missing"
    if status == "implemented" and len(text) < 50:
        return True, "stale"
    return False, None


def build_by_components_from_narrative(narrative: str) -> tuple:
    """
    Scan the narrative for mentions of known tools/services.
    Returns (by_components list, component_keys found).
    No control mapping — just name recognition.
    """
    component_keys = extract_components_from_text(narrative)
    by_components = []
    for key in component_keys:
        comp = KNOWN_COMPONENTS[key]
        by_components.append({
            "component-uuid": stable_uuid(f"component:{key}"),
            "uuid": stable_uuid(f"by-component:{key}:{narrative[:30]}"),
            "description": comp["title"],
            "implementation-status": {
                "state": "planned",
                "remarks": "referenced in SSP narrative — pending assessment verification"
            },
            "props": [
                {"name": "component-key", "value": key},
                {"name": "origin",        "value": "ssp-narrative"},
            ]
        })
    return by_components, component_keys


# ── Build component definitions ──────────────────────────────────────────────

def build_component_definitions(discovered_keys: set):
    """Build components section from what was actually found in narratives."""
    components = [
        {
            "uuid": stable_uuid("component:this-system"),
            "type": "this-system",
            "title": "Workshop Demo System",
            "description": (
                "A demo system built during the GRC Engineering Club "
                "OSCAL builder session."
            ),
            "status": {"state": "operational"}
        }
    ]
    for key in sorted(discovered_keys):
        comp = KNOWN_COMPONENTS[key]
        components.append({
            "uuid": stable_uuid(f"component:{key}"),
            "type": comp["type"],
            "title": comp["title"],
            "description": f"{comp['title']} — referenced in SSP control narratives.",
            "props": [
                {"name": "component-key", "value": key},
                {"name": "origin",        "value": "ssp-narrative"},
            ],
            "status": {"state": "operational"}
        })
    return components


# ── Build assessment results skeleton ────────────────────────────────────────

def build_assessment_results_skeleton(ssp_uuid: str):
    ar_uuid = stable_uuid("assessment-results:workshop")
    return {
        "assessment-results": {
            "uuid": ar_uuid,
            "metadata": {
                "title": "Workshop Demo — Assessment Results",
                "last-modified": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "version": "1.0.0",
                "oscal-version": "1.1.2",
                **oscal_roles_and_parties(),
            },
            "import-ap": {
                "href": "#",
                "remarks": "Assessment plan is automated — the pipeline scripts define what's checked and how."
            },
            "results": [
                {
                    "uuid": stable_uuid("result:workshop-run"),
                    "title": "Workshop Pipeline Run",
                    "description": "Evidence collected against FedRAMP Moderate baseline.",
                    "start": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "reviewed-controls": {
                        "control-selections": [
                            {"include-all": {}}
                        ]
                    },
                    "remarks": "Skeleton — populated by assess.py pipeline stage."
                }
            ]
        }
    }


# ── Build POA&M skeleton ─────────────────────────────────────────────────────

def build_poam_skeleton(ssp_uuid: str):
    return {
        "plan-of-action-and-milestones": {
            "uuid": stable_uuid("poam:workshop"),
            "metadata": {
                "title": "Workshop Demo — Plan of Action and Milestones",
                "last-modified": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "version": "1.0.0",
                "oscal-version": "1.1.2",
                **oscal_roles_and_parties(),
            },
            "import-ssp": {
                "href": "ssp.json"
            },
            "poam-items": [
                {
                    "uuid": stable_uuid("poam-item:placeholder"),
                    "title": "Placeholder — pending assessment",
                    "description": "This POA&M will be populated after the assessment and reconciliation stages run.",
                }
            ]
        }
    }


# ── Main converter ────────────────────────────────────────────────────────────

def convert_excel_to_oscal(input_path: str, output_dir: str):
    print(f"\n{'='*62}")
    print(f"  FedRAMP Moderate SSP → OSCAL Converter")
    print(f"  Discovery-Driven — SSP as claim, no assumptions")
    print(f"{'='*62}")
    print(f"  Input:      {input_path}")
    print(f"  Output dir: {output_dir}")
    print(f"{'='*62}\n")

    print(f"  Opening Excel SSP...")
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"  ERROR: Could not open Excel file: {e}")
        return False

    if SHEET_NAME not in wb.sheetnames:
        print(f"  ERROR: Sheet '{SHEET_NAME}' not found.")
        print(f"  Available sheets: {wb.sheetnames}")
        return False

    ws = wb[SHEET_NAME]
    print(f"  Found sheet: {SHEET_NAME}")
    print(f"  Extracting controls and scanning narratives for components...\n")

    implemented_requirements = []
    all_component_keys = set()
    stats = {
        "total": 0, "implemented": 0, "inherited": 0, "planned": 0,
        "not_applicable": 0, "has_narrative": 0, "missing_narrative": 0,
        "controls_with_components": 0, "controls_without_components": 0,
    }

    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        row_data = {key: get_cell(ws, row_num, col) for key, col in COLUMNS.items()}

        if not row_data["control_id"]:
            continue

        stats["total"] += 1
        control_id = normalize_control_id(row_data["control_id"])
        if not control_id:
            continue

        status = normalize_status(row_data["status"])
        family = control_id.split("-")[0] if "-" in control_id else ""
        ssp_text = row_data["implementation"]
        needs_review, review_reason = is_missing_or_stale(ssp_text, status)

        # Stats
        stat_map = {"implemented": "implemented", "inherited": "inherited",
                    "planned": "planned", "not-applicable": "not_applicable"}
        if status in stat_map:
            stats[stat_map[status]] += 1
        if ssp_text and not needs_review:
            stats["has_narrative"] += 1
        else:
            stats["missing_narrative"] += 1

        # Extract components from narrative
        by_components, component_keys = build_by_components_from_narrative(ssp_text or "")
        all_component_keys.update(component_keys)

        if component_keys:
            stats["controls_with_components"] += 1
        else:
            stats["controls_without_components"] += 1

        # Control-level props
        props = [
            {"name": "control-origination", "value": status},
            {"name": "control-family",       "value": family.upper()},
            {"name": "last-reconciled",      "value": "never"},
        ]

        if needs_review:
            props.append({"name": "review-flag", "value": review_reason})

        if ssp_text:
            props.append({"name": "baseline-narrative", "value": ssp_text})

        # Build the implemented-requirement
        statement = {
            "statement-id": f"{control_id}_smt",
            "uuid": stable_uuid(f"stmt:{control_id}"),
        }
        if ssp_text:
            statement["remarks"] = ssp_text
        if by_components:
            statement["by-components"] = by_components

        impl_req = {
            "uuid": stable_uuid(f"impl-req:{control_id}"),
            "control-id": control_id,
            "props": props,
            "statements": [statement],
        }
        implemented_requirements.append(impl_req)

        comp_display = ", ".join(component_keys) if component_keys else "(none)"
        print(f"  {control_id.upper():10s} {status:18s} components: {comp_display:30s} "
              f"{'✓' if ssp_text else '✗'} narrative")

    # ── Build the SSP ────────────────────────────────────────────────────────

    ssp_uuid = stable_uuid("ssp:workshop-demo")

    oscal_ssp = {
        "system-security-plan": {
            "uuid": ssp_uuid,
            "metadata": {
                "title": "Workshop Demo — System Security Plan",
                "last-modified": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "version": "1.0.0",
                "oscal-version": "1.1.2",
                **oscal_roles_and_parties(),
                "remarks": (
                    "Generated by GRC Engineering Club workshop converter. "
                    "FedRAMP Moderate baseline. Components extracted from SSP "
                    "narratives — discovery fills in the real inventory."
                ),
            },
            "import-profile": {
                "href": "https://raw.githubusercontent.com/usnistgov/oscal-content/main/nist.gov/SP800-53/rev5/json/NIST_SP-800-53_rev5_MODERATE-baseline_profile.json",
                "remarks": "FedRAMP Moderate baseline from NIST OSCAL content repository"
            },
            "system-characteristics": {
                "system-ids": [
                    {"id": "WORKSHOP-DEMO-001"}
                ],
                "system-name": "GRC Engineering Club Workshop Demo",
                "description": (
                    "A demonstration system built during the live builder session. "
                    "Components are extracted from SSP narratives. Discovery identifies "
                    "the actual environment inventory."
                ),
                "security-sensitivity-level": "moderate",
                "system-information": {
                    "information-types": [
                        {
                            "uuid": stable_uuid("info-type:demo"),
                            "title": "Workshop Demonstration Data",
                            "description": "Non-sensitive demonstration data for OSCAL pipeline training.",
                            "categorizations": [
                                {"system": "https://doi.org/10.6028/NIST.SP.800-60v2r1"}
                            ],
                            "confidentiality-impact": {"base": "fips-199-moderate"},
                            "integrity-impact":      {"base": "fips-199-moderate"},
                            "availability-impact":   {"base": "fips-199-moderate"},
                        }
                    ]
                },
                "security-impact-level": {
                    "security-objective-confidentiality": "fips-199-moderate",
                    "security-objective-integrity":       "fips-199-moderate",
                    "security-objective-availability":    "fips-199-moderate",
                },
                "status": {"state": "operational"},
                "authorization-boundary": {
                    "description": "AWS account, GitHub repositories, and associated security tooling."
                },
            },
            "system-implementation": {
                "users": [
                    {
                        "uuid": stable_uuid("user:workshop-participant"),
                        "title": "Workshop Participant",
                        "role-ids": ["system-owner"],
                        "description": "GRC Engineering Club member building the demo pipeline."
                    }
                ],
                "components": build_component_definitions(all_component_keys),
            },
            "control-implementation": {
                "description": (
                    "FedRAMP Moderate baseline controls with implementation narratives "
                    "as documented claims. Components extracted from narratives via "
                    "keyword matching. Discovery adds the real environment inventory."
                ),
                "implemented-requirements": implemented_requirements,
            }
        }
    }

    # ── Write all three OSCAL files ──────────────────────────────────────────

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    ssp_file = output_path / "ssp.json"
    with open(ssp_file, "w", encoding="utf-8") as f:
        json.dump(oscal_ssp, f, indent=2, ensure_ascii=False)

    ar = build_assessment_results_skeleton(ssp_uuid)
    ar_file = output_path / "assessment-results.json"
    with open(ar_file, "w", encoding="utf-8") as f:
        json.dump(ar, f, indent=2, ensure_ascii=False)

    poam = build_poam_skeleton(ssp_uuid)
    poam_file = output_path / "poam.json"
    with open(poam_file, "w", encoding="utf-8") as f:
        json.dump(poam, f, indent=2, ensure_ascii=False)

    # ── Summary ──────────────────────────────────────────────────────────────

    print(f"\n{'='*62}")
    print(f"  CONVERSION COMPLETE")
    print(f"{'='*62}")
    print(f"  OSCAL SSP:                {ssp_file}")
    print(f"  Assessment Results:       {ar_file}")
    print(f"  POA&M:                    {poam_file}")
    print(f"{'─'*62}")
    print(f"  Total controls:           {stats['total']}")
    print(f"  Implemented:              {stats['implemented']}")
    print(f"  Inherited:                {stats.get('inherited', 0)}")
    print(f"  Narratives present:       {stats['has_narrative']}")
    print(f"  Missing narratives:       {stats['missing_narrative']}")
    print(f"{'─'*62}")
    print(f"  Components from SSP:      {len(all_component_keys)}")
    if all_component_keys:
        print(f"    {', '.join(KNOWN_COMPONENTS[k]['title'] for k in sorted(all_component_keys))}")
    print(f"  Controls with components: {stats['controls_with_components']}")
    print(f"  Controls without:         {stats['controls_without_components']}")
    print(f"{'─'*62}")
    print(f"  UUID strategy:            v5 deterministic (stable diffs)")
    print(f"  Architecture:             SSP = claim | Discovery = truth | AR = evidence")
    print(f"{'='*62}")
    print(f"\n  Next: Run discovery to find what's actually in your environment")
    print(f"  python3 scripts/discover.py --ssp oscal/ssp.json --output oscal/inventory.json")
    print(f"{'='*62}\n")

    return True


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert FedRAMP Moderate template SSP (Excel) to OSCAL JSON"
    )
    parser.add_argument(
        "--input", "-i",
        default="Templates/fedramp-moderate-template-ssp.xlsx",
        help="Path to template SSP Excel file"
    )
    parser.add_argument(
        "--output", "-o",
        default="oscal",
        help="Output directory for OSCAL JSON files (default: oscal/)"
    )
    args = parser.parse_args()

    success = convert_excel_to_oscal(args.input, args.output)
    sys.exit(0 if success else 1)
