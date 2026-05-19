"""
export_poam.py — Export POA&M items to a formatted Excel report.

Reads oscal/poam.json and produces a styled .xlsx workbook.
"""

import sys, os, json, argparse
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(__file__))
from pipeline_utils import load_oscal

# ── Recommended-action mapping ──────────────────────────────────────────────
RECOMMENDED_ACTIONS = {
    "ia-2":  "Enable MFA for all interactive users",
    "sc-28": "Enable encryption at rest for all data stores",
    "cm-3":  "Enable branch protection on the default branch",
    "au-2":  "Enable VPC flow logs and centralized log collection",
    "au-9":  "Enable CloudTrail log file validation",
    "ac-2":  "Review and disable unused IAM accounts",
    "ac-6":  "Remove excessive IAM privileges (least privilege)",
    "sc-8":  "Enforce TLS/encryption in transit",
    "sc-12": "Rotate or establish cryptographic key management",
    "sa-11": "Enable static analysis (CodeQL / SAST) in CI pipeline",
    "si-2":  "Apply pending security patches and updates",
}

BANNER = """
==============================================================
  OSCAL Pipeline — Export POA&M Report
==============================================================
"""


def get_prop(props: list, name: str) -> str:
    """Return the value of the first prop matching *name*, or ''."""
    for p in props:
        if p.get("name") == name:
            return p.get("value", "")
    return ""


def recommended_action(control_id: str) -> str:
    return RECOMMENDED_ACTIONS.get(control_id.lower(), "Review and remediate")


def export_poam(poam_path: str, output_path: str):
    print(BANNER)

    data = load_oscal(poam_path)
    items = data.get("plan-of-action-and-milestones", {}).get("poam-items", [])
    print(f"  POA&M items found : {len(items)}")
    print(f"  Output file       : {output_path}")
    print()

    # ── Workbook setup ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POA&M Report"

    headers = [
        "POA&M ID",
        "Control ID",
        "Finding",
        "Source",
        "Description",
        "Status",
        "Date Found",
        "Recommended Action",
    ]

    # ── Styles ───────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    col_widths = {"A": 12, "B": 12, "C": 35, "D": 20, "E": 50, "F": 10, "G": 15, "H": 35}

    # ── Header row ───────────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, item in enumerate(items, start=2):
        props = item.get("props", [])
        control_id = get_prop(props, "control-id")
        source = get_prop(props, "source")
        status = get_prop(props, "status")
        created = get_prop(props, "created")

        # Parse created date for display
        date_found = ""
        if created:
            try:
                dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                date_found = dt.strftime("%Y-%m-%d")
            except ValueError:
                date_found = created

        seq = row_idx - 1
        values = [
            f"POAM-{seq:03d}",
            control_id.upper(),
            item.get("title", ""),
            source,
            item.get("description", ""),
            status,
            date_found,
            recommended_action(control_id),
        ]

        row_fill = white_fill if (seq % 2 == 1) else gray_fill

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ── Column widths ────────────────────────────────────────────────────────
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # ── Freeze & auto-filter ─────────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Save ─────────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"  Excel report saved: {output_path}")
    print()


def main():
    parser = argparse.ArgumentParser(description="Export POA&M to Excel")
    parser.add_argument("--poam", default="oscal/poam.json")
    parser.add_argument("--output", default="oscal/poam-report.xlsx")
    args = parser.parse_args()

    if not os.path.exists(args.poam):
        print(f"ERROR: POA&M file not found: {args.poam}")
        print("       Run the pipeline first to generate oscal/poam.json")
        sys.exit(1)

    export_poam(args.poam, args.output)


if __name__ == "__main__":
    main()
