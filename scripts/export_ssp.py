#!/usr/bin/env python3
"""
export_ssp.py — Export pipeline results back into an Excel SSP.

Reads the original SSP template and the pipeline's OSCAL JSON output,
produces an updated Excel file that reflects what the pipeline found.
"""

import sys, os, json, argparse
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.dirname(__file__))
from pipeline_utils import load_oscal

BANNER = """
==============================================================
  OSCAL Pipeline — Export Updated SSP
==============================================================
"""

# ── Verdict colours ────────────────────────────────────────────────────────────
VERDICT_FILLS = {
    "confirmed":     PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
    "contradicted":  PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # red
    "manual-review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # yellow
    "inherited":     PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),  # gray
}

BOLD = Font(bold=True)


def build_verdict_lookup(assessment: dict) -> dict:
    """
    Build control_id -> verdict from observation reconciliation-verdict props.
    If a control has mixed verdicts, the worst wins (contradicted > manual-review > confirmed).
    """
    results = assessment.get("assessment-results", {}).get("results", [])
    if not results:
        return {}

    result = results[0]
    obs_verdicts = {}  # control_id -> set of verdicts

    for obs in result.get("observations", []):
        control_id = None
        verdict = None
        for prop in obs.get("props", []):
            if prop["name"] == "control-id":
                control_id = prop["value"]
            if prop["name"] == "reconciliation-verdict":
                verdict = prop["value"]
        if control_id and verdict:
            obs_verdicts.setdefault(control_id, set()).add(verdict)

    # Collapse per-control verdict sets to a single verdict
    priority = {"contradicted": 0, "manual-review": 1, "confirmed": 2, "inherited": 3}
    verdicts = {}
    for cid, vset in obs_verdicts.items():
        best = sorted(vset, key=lambda v: priority.get(v, 99))[0]
        verdicts[cid] = best

    return verdicts


def build_evidence_method_lookup(assessment: dict) -> dict:
    """Build control_id -> evidence_method from findings props."""
    results = assessment.get("assessment-results", {}).get("results", [])
    if not results:
        return {}

    result = results[0]
    methods = {}
    for finding in result.get("findings", []):
        control_id = finding.get("target", {}).get("target-id")
        if not control_id:
            continue
        for prop in finding.get("props", []):
            if prop["name"] == "evidence-method":
                methods[control_id] = prop["value"]
                break
    return methods


def build_findings_summary(assessment: dict) -> dict:
    """Build control_id -> summary string of failed findings."""
    results = assessment.get("assessment-results", {}).get("results", [])
    if not results:
        return {}

    result = results[0]
    failures = {}  # control_id -> list of failure descriptions
    for finding in result.get("findings", []):
        target = finding.get("target", {})
        control_id = target.get("target-id")
        status = target.get("status", {}).get("state", "")
        if control_id and status == "not-satisfied":
            failures.setdefault(control_id, []).append(finding.get("title", ""))

    summaries = {}
    for cid, titles in failures.items():
        summaries[cid] = "; ".join(titles)
    return summaries


def build_component_lookup(ssp: dict) -> dict:
    """Build control_id -> list of component titles from SSP by-components."""
    impl_reqs = (
        ssp.get("system-security-plan", {})
        .get("control-implementation", {})
        .get("implemented-requirements", [])
    )
    components = {}
    for req in impl_reqs:
        control_id = req.get("control-id", "")
        comp_titles = set()
        for stmt in req.get("statements", []):
            for bc in stmt.get("by-components", []):
                desc = bc.get("description", "")
                if desc:
                    comp_titles.add(desc)
        if comp_titles:
            components[control_id] = sorted(comp_titles)
    return components


def build_ssp_origination_lookup(ssp: dict) -> dict:
    """Build control_id -> control-origination from SSP props."""
    impl_reqs = (
        ssp.get("system-security-plan", {})
        .get("control-implementation", {})
        .get("implemented-requirements", [])
    )
    originations = {}
    for req in impl_reqs:
        control_id = req.get("control-id", "")
        for prop in req.get("props", []):
            if prop["name"] == "control-origination":
                originations[control_id] = prop["value"]
                break
    return originations


def resolve_verdict(control_id_lower: str, verdicts: dict, findings_summary: dict,
                    originations: dict) -> str:
    """
    Determine the final verdict for a control.
    Uses observation reconciliation-verdict if available, otherwise infers.
    """
    if control_id_lower in verdicts:
        return verdicts[control_id_lower]

    # Check SSP origination for inherited
    if originations.get(control_id_lower) == "inherited":
        return "inherited"

    # Check findings
    if control_id_lower in findings_summary:
        return "contradicted"

    return "manual-review"


def auto_fit_columns(ws):
    """Approximate auto-fit for column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        # Cap width to avoid extremely wide columns
        adjusted = min(max_len + 2, 60)
        ws.column_dimensions[col_letter].width = adjusted


def main():
    parser = argparse.ArgumentParser(description="Export updated SSP to Excel")
    parser.add_argument("--ssp", default="oscal/ssp.json")
    parser.add_argument("--results", default="oscal/assessment-results.json")
    parser.add_argument("--template", default="Templates/fedramp-moderate-template-ssp.xlsx")
    parser.add_argument("--output", default="oscal/updated-ssp.xlsx")
    args = parser.parse_args()

    print(BANNER)

    # ── Validate inputs ────────────────────────────────────────────────────────
    for path, label in [(args.ssp, "SSP"), (args.results, "Assessment Results"),
                        (args.template, "Template")]:
        if not os.path.exists(path):
            print(f"  ERROR: {label} not found: {path}")
            sys.exit(1)

    # ── Load data ──────────────────────────────────────────────────────────────
    print("  Loading OSCAL data...")
    ssp = load_oscal(args.ssp)
    assessment = load_oscal(args.results)

    print("  Loading Excel template...")
    wb = openpyxl.load_workbook(args.template)
    ws = wb["FedRAMP Moderate Baseline"]

    # ── Build lookups ──────────────────────────────────────────────────────────
    verdicts = build_verdict_lookup(assessment)
    evidence_methods = build_evidence_method_lookup(assessment)
    findings_summary = build_findings_summary(assessment)
    component_lookup = build_component_lookup(ssp)
    originations = build_ssp_origination_lookup(ssp)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # ── Add new column headers in row 1 ────────────────────────────────────────
    new_headers = {
        10: "Reconciliation Verdict",
        11: "Last Reconciled",
        12: "Components Found",
        13: "Findings",
    }
    for col, header in new_headers.items():
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Also ensure original headers are bold
    for col in range(1, 10):
        ws.cell(row=1, column=col).font = BOLD

    # ── Process each control row ───────────────────────────────────────────────
    updated_count = 0
    verdict_counts = {"confirmed": 0, "contradicted": 0, "manual-review": 0, "inherited": 0}

    for row in range(2, ws.max_row + 1):
        control_id_raw = ws.cell(row=row, column=3).value
        if not control_id_raw:
            continue

        control_id_upper = str(control_id_raw).strip()
        control_id_lower = control_id_upper.lower()

        # Col 8: Evidence Method — update if we have assessment data
        if control_id_lower in evidence_methods:
            ws.cell(row=row, column=8, value=evidence_methods[control_id_lower].capitalize())

        # Determine verdict
        verdict = resolve_verdict(control_id_lower, verdicts, findings_summary, originations)

        # Col 10: Reconciliation Verdict
        verdict_cell = ws.cell(row=row, column=10, value=verdict)
        if verdict in VERDICT_FILLS:
            verdict_cell.fill = VERDICT_FILLS[verdict]
        verdict_cell.alignment = Alignment(horizontal="center")

        # Col 11: Last Reconciled
        ws.cell(row=row, column=11, value=today).alignment = Alignment(horizontal="center")

        # Col 12: Components Found
        components = component_lookup.get(control_id_lower, [])
        ws.cell(row=row, column=12, value=", ".join(components) if components else "")

        # Col 13: Findings
        findings_text = findings_summary.get(control_id_lower, "")
        ws.cell(row=row, column=13, value=findings_text)

        updated_count += 1
        verdict_counts[verdict] = verdict_counts.get(verdict, 0) + 1

    # ── Style and save ─────────────────────────────────────────────────────────
    auto_fit_columns(ws)

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    wb.save(args.output)

    # ── Summary ────────────────────────────────────────────────────────────────
    print(f"\n  Controls updated:  {updated_count}")
    print(f"  Verdicts:")
    for v in ["confirmed", "contradicted", "manual-review", "inherited"]:
        count = verdict_counts.get(v, 0)
        if count:
            print(f"    {v:20s} {count}")
    print(f"\n  Output: {args.output}")
    print()


if __name__ == "__main__":
    main()
